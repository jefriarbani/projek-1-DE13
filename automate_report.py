#Membuat Automated Reporting (Table & Grafik), dan dikirim ke discords

import pandas as pd #pandas untuk membuat dataframe(df)
from openpyxl import load_workbook #untuk berinterkasi antara python & excel file
from openpyxl.styles import *
from openpyxl.chart import *
from openpyxl.chart.shapes import GraphicalProperties
from openpyxl.chart.label import DataLabelList
from openpyxl.worksheet.dimensions import ColumnDimension, DimensionHolder
from openpyxl.utils import get_column_letter

input_file = 'input_data/supermarket_sales.xlsx'
output_file = 'data_output/report_penjualan_2019.xlsx'
# webhook_url = 'https://discordapp.com/api/webhooks/1111718857879605269/tkkwCuQNu2RKsg7kpGqr7l9a43DhRK9tUDjnq55yDwkKrbYcJY4D6tSh49VF3KzRxXtn'

##PART 1 - LOAD DATASET
df = pd.read_excel(input_file)
print (df.head())

#Penjualan Total per Gender & Product Line
df = df.pivot_table(index=['Gender','Date'],
                    columns='Product line', 
                    values='Total', 
                    aggfunc='sum').round()

print (df.head())
print('Save dataframe to excel...')

df.to_excel(output_file, 
                sheet_name='Report', 
                startrow=3)

print('Save dataframe done...')


#select active worksheet
wb = load_workbook(output_file)
wb.active = wb['Report']

min_column = wb.active.min_column
max_column = wb.active.max_column
min_row = wb.active.min_row
max_row = wb.active.max_row

print(min_column, max_column, min_row, max_row)

dim_holder = DimensionHolder(worksheet=wb.active)

for col in range(wb.active.min_column, wb.active.max_column + 1):
    dim_holder[get_column_letter(col)] = ColumnDimension(wb.active, min=col, max=col, width=20)

wb.active.column_dimensions = dim_holder


# ##PART 2 - GRAFIK

# barchart = BarChart()

barchart = BarChart()

data = Reference(wb.active, 
                min_col=min_column+2,
                max_col=max_column,
                min_row=min_row,
                max_row=max_row,
                )

categories = Reference(wb.active,
                        min_col=min_column,
                        max_col=min_column,
                        min_row=min_row+1,
                        max_row=max_row,
                        )

date = Reference(wb.active,
                        min_col=min_column+1,
                        max_col=min_column,
                        min_row=min_row+1,
                        max_row=max_row,
                        )

barchart.add_data(data, titles_from_data=True)
barchart.set_categories(categories)
barchart.set_categories(date)


wb.active.add_chart(barchart, 'j5')
barchart.title = 'Sales berdasarkan Produk'
barchart.style = 2
barchart.width = 100
barchart.height = 20
wb.save(output_file)


# #Total dari Penjualan
import string
alphabet = list(string.ascii_uppercase)
alphabet_excel = alphabet[:max_column]
#[A,B,C,D,E,F,G]
for i in alphabet_excel:
    if i != 'A':
        wb.active[f'{i}{max_row+1}'] = f'=SUM({i}{min_row+1}:{i}{max_row})'
        wb.active[f'{i}{max_row+1}'].style = 'Currency'

wb.active[f'{alphabet_excel[0]}{max_row+1}'] = 'Total'
wb.save(output_file)

wb.active['A1'] = 'Sales Report'
wb.active['A2'] = '2019'
wb.active['A1'].font = Font('Arial', bold=True, size=20)
wb.active['A2'].font = Font('Arial', bold=True, size=10)

wb.save(output_file)

# #PART - 3 Send to discord
# # pip3 install discord==1.7.3

# def send_to_discord():
#     import discord
#     from discord import SyncWebhook

#     webhook = SyncWebhook.from_url(webhook_url)

#     with open(file=output_file, mode='rb') as file:
#         excel_file = discord.File(file)

#     webhook.send('This is an automated report', 
#                 username='Sales Bot', 
#                 file=excel_file)

# send_to_discord()

